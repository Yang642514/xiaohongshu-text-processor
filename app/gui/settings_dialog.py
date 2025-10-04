from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QFormLayout, QLineEdit, QPushButton, QFileDialog, QMessageBox, QHBoxLayout, QWidget, QSizePolicy
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from .message_dialog import MessageDialog


class SettingsDialog(QDialog):
    def __init__(self, settings: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("设置")
        # 使用共享设置对象（只读回显；返回 get_settings 由主窗口写入保存）
        self.settings = settings

        layout = QVBoxLayout()
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form.setFormAlignment(Qt.AlignLeft)
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(6)
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        form.setRowWrapPolicy(QFormLayout.DontWrapRows)

        # 无图模板路径
        self.tpl_edit = QLineEdit(self.settings.get("template_excel_path", ""))
        self.tpl_edit.setFixedHeight(26)
        self.tpl_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        # 有图模板路径
        self.tpl_img_edit = QLineEdit(self.settings.get("template_excel_image_path", ""))
        self.tpl_img_edit.setFixedHeight(26)
        self.tpl_img_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.zip_edit = QLineEdit(self.settings.get("zip_output_dir", "output"))
        self.zip_edit.setFixedHeight(26)
        self.zip_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.page_col = QLineEdit(self.settings.get("page_column", "页面"))
        self.point_title_col = QLineEdit(self.settings.get("point_title_column", "文本_1"))
        self.point_content_col = QLineEdit(self.settings.get("point_content_column", "文本_2"))
        # 新增：与默认内容
        self.extra_text_col = QLineEdit(self.settings.get("extra_text_column", "文本_3"))
        self.extra_text_default = QLineEdit(self.settings.get("extra_text_default", "内容仅供参考，身体不适请及时就医!"))
        # 图片列名（用于配图版模板写入）
        self.image_col = QLineEdit(self.settings.get("image_column", "图片_1"))
        # 以下逻辑已内置：再次处理不嵌套与尾段过滤始终启用；不在设置中展示。

        # 模板路径（右侧内联浏览按钮）
        tpl_row = QHBoxLayout()
        tpl_row.setContentsMargins(0, 0, 0, 0)
        tpl_row.setSpacing(8)
        tpl_row.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        tpl_row.addWidget(self.tpl_edit)
        btn_browse_tpl = QPushButton("浏览")
        btn_browse_tpl.setObjectName("InlineButton")
        btn_browse_tpl.setFixedHeight(24)
        try:
            btn_browse_tpl.setMinimumWidth(64)
        except Exception:
            pass
        btn_browse_tpl.clicked.connect(self._browse_template)
        tpl_row.addWidget(btn_browse_tpl)
        # 确保编辑框在该行中占据更多空间，避免显示不完整
        tpl_row.setStretch(0, 1)
        tpl_row.setStretch(1, 0)
        tpl_container = QWidget()
        tpl_container.setLayout(tpl_row)
        tpl_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        # 使用自定义标签以控制高度与垂直对齐
        form.addRow(self._vcenter_label("无图模板路径"), tpl_container)

        # 有图模板路径（右侧内联浏览按钮）
        tpl_img_row = QHBoxLayout()
        tpl_img_row.setContentsMargins(0, 0, 0, 0)
        tpl_img_row.setSpacing(8)
        tpl_img_row.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        tpl_img_row.addWidget(self.tpl_img_edit)
        btn_browse_tpl_img = QPushButton("浏览")
        btn_browse_tpl_img.setObjectName("InlineButton")
        btn_browse_tpl_img.setFixedHeight(24)
        try:
            btn_browse_tpl_img.setMinimumWidth(64)
        except Exception:
            pass
        btn_browse_tpl_img.clicked.connect(self._browse_template_image)
        tpl_img_row.addWidget(btn_browse_tpl_img)
        tpl_img_row.setStretch(0, 1)
        tpl_img_row.setStretch(1, 0)
        tpl_img_container = QWidget()
        tpl_img_container.setLayout(tpl_img_row)
        tpl_img_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        form.addRow(self._vcenter_label("有图模板路径"), tpl_img_container)

        # 压缩包目录（右侧内联浏览按钮）
        zip_row = QHBoxLayout()
        zip_row.setContentsMargins(0, 0, 0, 0)
        zip_row.setSpacing(8)
        zip_row.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        zip_row.addWidget(self.zip_edit)
        btn_browse_zip = QPushButton("浏览")
        btn_browse_zip.setObjectName("InlineButton")
        btn_browse_zip.setFixedHeight(24)
        try:
            btn_browse_zip.setMinimumWidth(64)
        except Exception:
            pass
        btn_browse_zip.clicked.connect(self._browse_zip_dir)
        zip_row.addWidget(btn_browse_zip)
        # 同样设置伸展，保证前两个编辑框完整显示
        zip_row.setStretch(0, 1)
        zip_row.setStretch(1, 0)
        zip_container = QWidget()
        zip_container.setLayout(zip_row)
        zip_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        form.addRow(self._vcenter_label("压缩包目录"), zip_container)

        form.addRow(self._vcenter_label("页面列名"), self.page_col)
        form.addRow(self._vcenter_label("分论点标题列名"), self.point_title_col)
        form.addRow(self._vcenter_label("分论点内容列名"), self.point_content_col)
        form.addRow(self._vcenter_label("页面注释列名"), self.extra_text_col)
        form.addRow(self._vcenter_label("页面注释内容"), self.extra_text_default)
        form.addRow(self._vcenter_label("图片列名"), self.image_col)

        # （移除）飞书配置入口，迁移至待同步页面
        # 不展示：再次处理不嵌套、尾段过滤开关，以及尾段关键词相关配置
        
        layout.addLayout(form)

        # 底部仅保留“检测模板”和“保存”按钮，并适当增大高度与右对齐
        btns = QHBoxLayout()
        btns.setSpacing(14)
        btns.setAlignment(Qt.AlignCenter)
        btn_check = QPushButton("检测无图模板可用性")
        btn_check.setFixedHeight(30)
        btn_check.clicked.connect(self._check_template)
        # 含图片列的模板检测
        btn_check_img = QPushButton("检测有图模板（含图片列）")
        btn_check_img.setFixedHeight(30)
        btn_check_img.clicked.connect(self._check_template_with_image)
        btn_ok = QPushButton("保存")
        btn_ok.setFixedHeight(30)
        btn_ok.clicked.connect(self.accept)
        btns.addWidget(btn_check)
        btns.addWidget(btn_check_img)
        btns.addWidget(btn_ok)
        layout.addLayout(btns)

        self.setLayout(layout)
        # 默认更宽并居中显示，但初始位置上移一些
        self.resize(540, 360)
        self.setModal(True)
        self.setWindowModality(Qt.ApplicationModal)
        # 居中基础上向上偏移 80 像素（可按需调整）
        parent_geom = self.parent().geometry() if self.parent() else self.geometry()
        offset_y = 80
        self.setGeometry(
            parent_geom.center().x() - self.width() // 2,
            parent_geom.center().y() - self.height() // 2 - offset_y,
            self.width(), self.height()
        )

    def _vcenter_label(self, text: str):
        from PyQt5.QtWidgets import QLabel
        lbl = QLabel(text)
        lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        # 固定高度以与右侧编辑框对齐（更紧凑）
        lbl.setFixedHeight(26)
        lbl.setFixedWidth(110)
        return lbl

    def _browse_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "模板Excel", "", "Excel Files (*.xlsx)")
        if path:
            self.tpl_edit.setText(path)

    def _browse_template_image(self):
        path, _ = QFileDialog.getOpenFileName(self, "有图模板Excel", "", "Excel Files (*.xlsx)")
        if path:
            self.tpl_img_edit.setText(path)

    def _browse_zip_dir(self):
        path = QFileDialog.getExistingDirectory(self, "压缩包目录", "")
        if path:
            self.zip_edit.setText(path)

    def get_settings(self) -> dict:
        return {
            "template_excel_path": self.tpl_edit.text().strip(),
            "template_excel_image_path": self.tpl_img_edit.text().strip(),
            "zip_output_dir": self.zip_edit.text().strip(),
            "page_column": self.page_col.text().strip() or "页面",
            "point_title_column": self.point_title_col.text().strip() or "文本_1",
            "point_content_column": self.point_content_col.text().strip() or "文本_2",
            "extra_text_column": self.extra_text_col.text().strip() or "文本_3",
            "extra_text_default": self.extra_text_default.text().strip() or "内容仅供参考，身体不适请及时就医!",
            "image_column": self.image_col.text().strip() or "图片_1",
            # （移除）飞书相关配置由独立配置页负责保存
        }

    def _check_template(self):
        path = self.tpl_edit.text().strip()
        if not path:
            MessageDialog.warning(self, "提示", "请先选择模板Excel路径")
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            need_cols = [
                self.page_col.text().strip() or "页面",
                self.point_title_col.text().strip() or "文本_1",
                self.point_content_col.text().strip() or "文本_2",
                self.extra_text_col.text().strip() or "文本_3",
            ]
            header_row = None
            for r in range(1, min(10, ws.max_row) + 1):
                row_vals = [c.value for c in ws[r]]
                if row_vals and all(col in row_vals for col in need_cols):
                    header_row = r
                    break
            if header_row is None:
                MessageDialog.error(self, "模板不可用", f"未在前10行找到包含列：{', '.join(need_cols)} 的表头")
            else:
                MessageDialog.info(self, "模板可用", f"模板检测通过（表头行：第{header_row}行）")
        except Exception as e:
            MessageDialog.error(self, "错误", f"无法读取模板：{e}")

    def _check_template_with_image(self):
        """在基本列的基础上，要求包含图片列名（默认 图片_1）。"""
        path = self.tpl_img_edit.text().strip()
        if not path:
            MessageDialog.warning(self, "提示", "请先选择有图模板Excel路径")
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            need_cols = [
                self.page_col.text().strip() or "页面",
                self.point_title_col.text().strip() or "文本_1",
                self.point_content_col.text().strip() or "文本_2",
                self.extra_text_col.text().strip() or "文本_3",
                self.image_col.text().strip() or "图片_1",
            ]
            header_row = None
            for r in range(1, min(10, ws.max_row) + 1):
                row_vals = [c.value for c in ws[r]]
                if row_vals and all(col in row_vals for col in need_cols):
                    header_row = r
                    break
            if header_row is None:
                MessageDialog.error(self, "模板不可用", f"未在前10行找到包含列：{', '.join(need_cols)} 的表头")
            else:
                MessageDialog.info(self, "模板可用", f"有图模板检测通过（包含图片列；表头行：第{header_row}行）")
        except Exception as e:
            MessageDialog.error(self, "错误", f"无法读取模板：{e}")