import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook


def write_to_template(template_path: str, output_dir: str, title: str,
                      points: List[Tuple[str, str]],
                      column_map: Dict[str, str]) -> str:
    """
    将解析结果写入模板 Excel 的下一行：
    - column_map: { title_column, point_column, content_column }
    返回生成的输出文件路径。
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    # 读取表头行（第一行）以定位列索引
    # 扫描前10行以定位表头行（可能不在第1行）
    page_col_name = column_map.get("page_column", "页面") or "页面"
    point_title_col_name = column_map.get("point_title_column", "文本_1") or "文本_1"
    point_content_col_name = column_map.get("point_content_column", "文本_2") or "文本_2"
    # 新增必填列：文本_3（列名可配置），写入固定默认内容
    extra_text_col_name = column_map.get("extra_text_column", "文本_3") or "文本_3"
    extra_text_default = column_map.get("extra_text_default", "内容仅供参考，身体不适请及时就医!")
    need_cols = [page_col_name, point_title_col_name, point_content_col_name, extra_text_col_name]

    header_row = None
    headers: List[Optional[str]] = []
    for r in range(1, min(10, ws.max_row) + 1):
        row_vals = [c.value for c in ws[r]]
        if row_vals and all(col in row_vals for col in need_cols):
            header_row = r
            headers = row_vals
            break

    if header_row is None:
        # 汇总候选行信息以便用户排查
        raise ValueError(
            f"未在前10行找到表头，需包含列：{', '.join(need_cols)}"
        )

    def try_col_index_from_headers(name: Optional[str]) -> Optional[int]:
        if not name:
            return None
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    page_col = try_col_index_from_headers(page_col_name)
    point_title_col = try_col_index_from_headers(point_title_col_name)
    point_content_col = try_col_index_from_headers(point_content_col_name)
    extra_text_col = try_col_index_from_headers(extra_text_col_name)

    if page_col is None or point_title_col is None or point_content_col is None or extra_text_col is None:
        missing = []
        if page_col is None:
            missing.append(page_col_name)
        if point_title_col is None:
            missing.append(point_title_col_name)
        if point_content_col is None:
            missing.append(point_content_col_name)
        if extra_text_col is None:
            missing.append(extra_text_col_name)
        raise ValueError(f"模板缺少必需列: {', '.join(missing)}（表头行: 第{header_row}行）")

    # 定位写入起始行（在现有数据之后的下一行）
    write_row_start = ws.max_row + 1

    # 写入多行：每个分论点一行，包含 页面 / 标题 / 内容
    numerals = ["一","二","三","四","五","六","七","八","九","十","十一","十二","十三","十四","十五","十六","十七","十八","十九","二十"]
    for idx, (pt_title, pt_content) in enumerate(points):
        row = write_row_start + idx
        num_cn = numerals[idx] if idx < len(numerals) else f"{idx+1}"
        page_val = f"分论点{num_cn}页面"
        ws.cell(row=row, column=page_col, value=page_val)
        ws.cell(row=row, column=point_title_col, value=pt_title)
        # 直接写入内容，不做任何删改
        ws.cell(row=row, column=point_content_col, value=pt_content)
        # 写入文本_3默认内容
        ws.cell(row=row, column=extra_text_col, value=extra_text_default)

    # 输出文件名：仅标题.xlsx（同名覆盖）
    os.makedirs(output_dir, exist_ok=True)
    safe_title = "".join(ch for ch in title if ch not in "\\/:*?\"<>|") or "输出"
    out_path = os.path.join(output_dir, f"{safe_title}.xlsx")
    wb.save(out_path)
    return out_path